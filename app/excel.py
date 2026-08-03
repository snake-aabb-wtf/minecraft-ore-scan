import math
from pathlib import Path
from openpyxl import Workbook
from .constants import ORE_OPTIONS


def export_to_excel(positions, output_path: Path, ox: int, oy: int, oz: int, log_callback=None):
    ore_name_map = dict(ORE_OPTIONS)
    rows = []
    for x, y, z, block_id in positions:
        dist_sq = (x - ox) ** 2 + (y - oy) ** 2 + (z - oz) ** 2
        display = ore_name_map.get(block_id, block_id)
        rows.append((dist_sq, x, y, z, display))
    rows.sort(key=lambda r: (r[0], r[2], r[1], r[3]))

    if log_callback:
        log_callback("正在写入 Excel...")
    MAX_ROWS = 1_048_575
    wb = Workbook(write_only=True)
    sheet_num = 0
    sheet = None
    row_count = MAX_ROWS
    headers = ["X", "Y", "Z", "Mineral", f"DistanceTo_({ox},{oy},{oz})"]
    for dist_sq, x, y, z, name in rows:
        if row_count >= MAX_ROWS:
            sheet_num += 1
            sheet = wb.create_sheet(f"Minerals_{sheet_num}")
            sheet.append(headers)
            row_count = 0
        sheet.append([x, y, z, name, math.sqrt(dist_sq)])
        row_count += 1
    if sheet is None:
        sheet = wb.create_sheet("Minerals_1")
        sheet.append(headers)
        sheet_num = 1
    wb.save(output_path)
    return len(rows), sheet_num
