"""Scan Minecraft Anvil chunks for block IDs and write distance-sorted Excel rows."""

import argparse
import gzip
import io
import math
import struct
import sys
import zlib
from pathlib import Path

import nbtlib
from openpyxl import Workbook


DIMENSION_PATHS = {
    "minecraft:overworld": ("region",),
    "minecraft:the_nether": ("DIM-1", "region"),
    "minecraft:the_end": ("DIM1", "region"),
}
MAX_EXCEL_DATA_ROWS = 1_048_575


def region_path(world, dimension, region_x, region_z):
    return world.joinpath(*DIMENSION_PATHS[dimension], f"r.{region_x}.{region_z}.mca")


def unpack_block_indices(data, palette_size):
    bits = max(4, (palette_size - 1).bit_length())
    mask = (1 << bits) - 1
    values_per_long = 64 // bits
    values = []
    for index in range(4096):
        long_index, slot = divmod(index, values_per_long)
        offset = slot * bits
        value = (data[long_index] & ((1 << 64) - 1)) >> offset
        values.append(value & mask)
    return values


def read_chunk_nbt(data):
    return nbtlib.File.parse(io.BytesIO(data))


def scan_chunk(data, chunk_x, chunk_z, target_ids):
    root = read_chunk_nbt(data)
    positions = []
    for section in root.get("sections", []):
        block_states = section.get("block_states")
        if not block_states:
            continue
        palette = block_states.get("palette", [])
        target_indices = {
            index for index, state in enumerate(palette)
            if str(state.get("Name")) in target_ids
        }
        if not target_indices:
            continue
        section_y = int(section.get("Y", 0))
        packed = block_states.get("data")
        if packed is None:
            indices = [0] * 4096
        else:
            indices = unpack_block_indices([int(value) for value in packed], len(palette))
        for index, palette_index in enumerate(indices):
            if palette_index not in target_indices:
                continue
            local_x = index & 15
            local_z = (index >> 4) & 15
            local_y = (index >> 8) & 15
            positions.append((
                chunk_x * 16 + local_x,
                section_y * 16 + local_y,
                chunk_z * 16 + local_z,
                str(palette[palette_index]["Name"]),
            ))
    return positions


def read_chunk_payload(region, sector_offset):
    region.seek(sector_offset * 4096)
    length_data = region.read(4)
    if len(length_data) != 4:
        return None
    length = struct.unpack(">i", length_data)[0]
    compression = region.read(1)[0]
    payload = region.read(length - 1)
    if compression == 1:
        return gzip.decompress(payload)
    if compression == 2:
        return zlib.decompress(payload)
    if compression == 3:
        return payload
    raise ValueError(f"unsupported chunk compression type: {compression}")


def scan_region(path, min_chunk, max_chunk, target_ids, positions):
    parts = path.stem.split(".")
    region_x, region_z = int(parts[1]), int(parts[2])
    scanned = 0
    with path.open("rb") as region:
        header = region.read(4096)
        if len(header) != 4096:
            return 0
        for local_z in range(32):
            chunk_z = region_z * 32 + local_z
            if not min_chunk <= chunk_z <= max_chunk:
                continue
            for local_x in range(32):
                chunk_x = region_x * 32 + local_x
                if not min_chunk <= chunk_x <= max_chunk:
                    continue
                entry = 4 * (local_x + local_z * 32)
                sector_offset = (header[entry] << 16) | (header[entry + 1] << 8) | header[entry + 2]
                if sector_offset == 0:
                    continue
                payload = read_chunk_payload(region, sector_offset)
                if payload:
                    positions.extend(scan_chunk(payload, chunk_x, chunk_z, target_ids))
                    scanned += 1
    return scanned


def parse_ore_spec(spec):
    if "=" in spec:
        block_id, display_name = spec.split("=", 1)
    else:
        block_id, display_name = spec, spec
    block_id = block_id.strip()
    display_name = display_name.strip()
    if not block_id or not display_name:
        raise ValueError(f"invalid --ore value: {spec!r}")
    if ":" not in block_id:
        block_id = f"minecraft:{block_id}"
    return block_id, display_name


def parse_origin(value):
    values = [int(part.strip()) for part in value.split(",")]
    if len(values) != 3:
        raise ValueError("--origin must be X,Y,Z")
    return tuple(values)


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--world", type=Path, default=Path("world"))
    parser.add_argument("--dimension", choices=sorted(DIMENSION_PATHS), default="minecraft:overworld")
    parser.add_argument("--min-chunk", type=int, required=True)
    parser.add_argument("--max-chunk", type=int, required=True)
    parser.add_argument("--ore", action="append", required=True, metavar="BLOCK_ID=中文名")
    parser.add_argument("--origin", default="0,64,0", help="distance origin as X,Y,Z")
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def main():
    args = parse_args()
    if args.min_chunk > args.max_chunk:
        raise ValueError("--min-chunk must be <= --max-chunk")
    target_names = dict(parse_ore_spec(spec) for spec in args.ore)
    target_ids = set(target_names)
    origin_x, origin_y, origin_z = parse_origin(args.origin)
    positions = []
    scanned = 0
    regions = 0

    region_min = args.min_chunk // 32
    region_max = args.max_chunk // 32
    for region_z in range(region_min, region_max + 1):
        for region_x in range(region_min, region_max + 1):
            path = region_path(args.world, args.dimension, region_x, region_z)
            if not path.exists():
                continue
            regions += 1
            scanned += scan_region(path, args.min_chunk, args.max_chunk, target_ids, positions)
            print(f"scanned regions={regions}, chunks={scanned}, blocks={len(positions)}", flush=True)

    rows = []
    for x, y, z, block_id in positions:
        distance_squared = (x - origin_x) ** 2 + (y - origin_y) ** 2 + (z - origin_z) ** 2
        rows.append((distance_squared, x, y, z, target_names[block_id]))
    rows.sort(key=lambda row: (row[0], row[2], row[1], row[3]))

    workbook = Workbook(write_only=True)
    sheet_number = 0
    sheet = None
    row_in_sheet = MAX_EXCEL_DATA_ROWS
    headers = ["X", "Y", "Z", "Mineral", f"DistanceTo_({origin_x},{origin_y},{origin_z})"]
    for distance_squared, x, y, z, display_name in rows:
        if row_in_sheet >= MAX_EXCEL_DATA_ROWS:
            sheet_number += 1
            sheet = workbook.create_sheet(f"Minerals_{sheet_number}")
            sheet.append(headers)
            row_in_sheet = 0
        sheet.append([x, y, z, display_name, math.sqrt(distance_squared)])
        row_in_sheet += 1

    if sheet is None:
        sheet_number = 1
        sheet = workbook.create_sheet("Minerals_1")
        sheet.append(headers)
    workbook.save(args.output)
    print(f"scanned_chunks={scanned} blocks={len(rows)} sheets={sheet_number} output={args.output}", flush=True)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr, flush=True)
        raise
