"""app.excel 导出器的单元测试：排序、零结果边界、可重新打开、显示名映射。

跨 worksheet 边界（>1,048,575 行）需要写入百万行，速度与资源代价过高，
不在单元测试中覆盖；该边界由产品代码的 MAX_ROWS 逻辑保证。
"""
import math
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.excel import export_to_excel


def _read_rows(path: Path):
    """读回导出文件，返回 (sheetnames, rows) 列表。"""
    workbook = load_workbook(path, read_only=True)
    try:
        sheetnames = workbook.sheetnames
        rows = []
        for name in sheetnames:
            sheet_rows = list(workbook[name].iter_rows(values_only=True))
            if sheet_rows:
                rows.append(sheet_rows)
        return sheetnames, rows
    finally:
        workbook.close()


class ExportToExcelTest(unittest.TestCase):
    def test_empty_positions_creates_minerals_1_and_returns_one_sheet(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "empty.xlsx"
            rows_count, sheets_count = export_to_excel([], output_path, 0, 0, 0)
            self.assertEqual(rows_count, 0)
            # 文件实际创建了含表头的 Minerals_1，返回统计必须与之一致
            self.assertEqual(sheets_count, 1)

            sheetnames, sheet_rows = _read_rows(output_path)
            self.assertEqual(sheetnames, ["Minerals_1"])
            self.assertEqual(len(sheet_rows), 1)
            headers = sheet_rows[0][0]
            self.assertEqual(headers, ("X", "Y", "Z", "Mineral", "DistanceTo_(0,0,0)"))

    def test_single_position(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "single.xlsx"
            rows_count, sheets_count = export_to_excel(
                [(10, 20, 30, "minecraft:stone")], output_path, 0, 0, 0
            )
            self.assertEqual((rows_count, sheets_count), (1, 1))

            _, sheet_rows = _read_rows(output_path)
            data = sheet_rows[0][1]
            expected_dist = math.sqrt(10 ** 2 + 20 ** 2 + 30 ** 2)
            self.assertEqual(data[:4], (10, 20, 30, "minecraft:stone"))
            self.assertAlmostEqual(data[4], expected_dist, places=6)

    def test_ore_display_name_uses_chinese_name(self):
        # minecraft:diamond_ore 在 ORE_OPTIONS 中的显示名是 "钻石"
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "name.xlsx"
            export_to_excel([(1, 2, 3, "minecraft:diamond_ore")], output_path, 0, 64, 0)
            _, sheet_rows = _read_rows(output_path)
            self.assertEqual(sheet_rows[0][1][3], "钻石")

    def test_unknown_block_id_falls_back_to_raw_id(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "unknown.xlsx"
            export_to_excel([(1, 2, 3, "minecraft:not_an_ore")], output_path, 0, 64, 0)
            _, sheet_rows = _read_rows(output_path)
            self.assertEqual(sheet_rows[0][1][3], "minecraft:not_an_ore")

    def test_sorted_by_distance_then_y_then_x_then_z(self):
        # 排序键为 (dist_sq, y, x, z)。距原点 (0,0,0) 平方距离为 1 的三个点：
        #   (0,0,1) -> y=0, x=0
        #   (1,0,0) -> y=0, x=1
        #   (0,1,0) -> y=1
        # 因此顺序为 (0,0,1) < (1,0,0) < (0,1,0)（先比 y 再比 x 再比 z）。
        positions = [
            (0, 0, 1, "minecraft:stone"),
            (5, 5, 5, "minecraft:stone"),
            (0, 1, 0, "minecraft:stone"),
            (1, 0, 0, "minecraft:stone"),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "sorted.xlsx"
            export_to_excel(positions, output_path, 0, 0, 0)
            _, sheet_rows = _read_rows(output_path)
            coords = [row[0:3] for row in sheet_rows[0][1:]]
            self.assertEqual(coords, [(0, 0, 1), (1, 0, 0), (0, 1, 0), (5, 5, 5)])

    def test_distance_header_reflects_origin(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "origin.xlsx"
            export_to_excel([(1, 2, 3, "minecraft:stone")], output_path, 10, -5, 7)
            _, sheet_rows = _read_rows(output_path)
            headers = sheet_rows[0][0]
            self.assertEqual(headers[-1], "DistanceTo_(10,-5,7)")


if __name__ == "__main__":
    unittest.main()
