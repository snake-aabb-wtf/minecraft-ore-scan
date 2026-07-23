import hashlib
import os
from collections import deque
from pathlib import Path
import shutil
import socket
import tempfile
import time
import unittest

from openpyxl import load_workbook

from app.constants import RCON_PASSWORD
from app.excel import export_to_excel
from app.installer import (
    download_server,
    fetch_manifest,
    fetch_version_data,
    get_server_java_requirement,
    update_server_properties,
)
from app.java_runtime import find_java_runtimes, select_java_runtime
from app.rcon import RconClient
from app.world import get_region_dir, location_entry_exists, pregenerate_chunks, scan_all_regions, start_server


SMOKE_VERSION = "1.20.5"
SMOKE_JAVA_MAJOR = 21
SMOKE_CHUNK = 32
SERVER_START_TIMEOUT_SECONDS = 180
SERVER_STOP_TIMEOUT_SECONDS = 60


def _free_port():
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return sock.getsockname()[1]


def _update_properties(path: Path, updates: dict):
    properties = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if "=" in line and not line.startswith("#"):
            key, value = line.split("=", 1)
            properties[key] = value
    properties.update(updates)
    path.write_text(
        "".join(f"{key}={value}\n" for key, value in properties.items()),
        encoding="utf-8",
    )


@unittest.skipUnless(
    os.environ.get("RUN_MINECRAFT_E2E") == "1",
    "set RUN_MINECRAFT_E2E=1 to run the real Minecraft server smoke test",
)
class MinecraftEndToEndSmokeTest(unittest.TestCase):
    def _wait_for_rcon(self, process, port: int):
        deadline = time.monotonic() + SERVER_START_TIMEOUT_SECONDS
        last_error = None
        while time.monotonic() < deadline:
            if process.poll() is not None:
                raise RuntimeError(f"Minecraft server exited with code {process.returncode}")
            try:
                client = RconClient("127.0.0.1", port, RCON_PASSWORD, timeout=5)
                client.command("list", retries=0)
                return client
            except Exception as exc:
                last_error = exc
                time.sleep(2)
        raise RuntimeError(f"Minecraft server did not accept RCON connections: {last_error}")

    def _stop_server(self, process, rcon):
        if rcon:
            try:
                rcon.command("stop", retries=0)
            except Exception:
                pass
            rcon.close()
        if process and process.poll() is None:
            try:
                process.wait(timeout=SERVER_STOP_TIMEOUT_SECONDS)
            except Exception:
                process.kill()
                process.wait(timeout=15)
        if process:
            for stream in (process.stdin, process.stdout, process.stderr):
                if stream:
                    stream.close()

    def _write_diagnostics(self, server_dir: Path, server_output):
        artifacts_dir = os.environ.get("CI_SMOKE_ARTIFACT_DIR")
        if not artifacts_dir:
            return

        destination = Path(artifacts_dir)
        destination.mkdir(parents=True, exist_ok=True)
        (destination / "server-output.log").write_text(
            "\n".join(server_output) + "\n",
            encoding="utf-8",
        )
        for relative_path in (Path("logs") / "latest.log", Path("server.log")):
            source = server_dir / relative_path
            if source.exists():
                shutil.copy2(source, destination / source.name)

    def test_download_start_generate_scan_and_export(self):
        manifest = fetch_manifest()
        server_output = deque(maxlen=250)
        process = None
        rcon = None

        with tempfile.TemporaryDirectory(prefix="minecraft-ore-scan-e2e-") as temp_dir:
            server_dir = Path(temp_dir) / SMOKE_VERSION
            try:
                downloaded_size = download_server(server_dir, SMOKE_VERSION, manifest)
                server_jar = server_dir / "server.jar"
                version_data = fetch_version_data(SMOKE_VERSION, manifest)
                expected_sha1 = version_data["downloads"]["server"]["sha1"]

                self.assertGreater(downloaded_size, 0)
                self.assertTrue(server_jar.exists())
                self.assertEqual(
                    hashlib.sha1(server_jar.read_bytes()).hexdigest(),
                    expected_sha1,
                )

                requirement = get_server_java_requirement(server_dir)
                self.assertEqual(requirement.major_version, SMOKE_JAVA_MAJOR)
                runtime = select_java_runtime(requirement.major_version, find_java_runtimes())
                self.assertIsNotNone(runtime, f"Java {requirement.major_version} was not discovered")

                update_server_properties(server_dir / "server.properties", "123456789")
                server_port = _free_port()
                rcon_port = _free_port()
                while rcon_port == server_port:
                    rcon_port = _free_port()
                _update_properties(
                    server_dir / "server.properties",
                    {
                        "server-port": str(server_port),
                        "rcon.port": str(rcon_port),
                    },
                )

                process = start_server(
                    server_dir,
                    java_executable=runtime.executable,
                    log_callback=server_output.append,
                )
                rcon = self._wait_for_rcon(process, rcon_port)
                rcon.command("seed", retries=0)

                world_dir = server_dir / "world"
                region_dir = get_region_dir(world_dir, "minecraft:overworld")
                self.assertTrue(
                    pregenerate_chunks(
                        rcon,
                        "minecraft:overworld",
                        SMOKE_CHUNK,
                        SMOKE_CHUNK,
                        region_dir,
                        running_check=lambda: True,
                    )
                )
                self.assertTrue(location_entry_exists(region_dir, SMOKE_CHUNK, SMOKE_CHUNK))

                rcon.command("save-all flush", retries=1)
                rcon.command("stop", retries=1)
                rcon.close()
                rcon = None
                process.wait(timeout=SERVER_STOP_TIMEOUT_SECONDS)

                positions = scan_all_regions(
                    region_dir,
                    SMOKE_CHUNK,
                    SMOKE_CHUNK,
                    {"minecraft:stone"},
                    running_check=lambda: True,
                )
                self.assertTrue(positions, "generated chunk did not contain scanable stone blocks")

                output_path = Path(temp_dir) / "smoke-results.xlsx"
                rows_count, sheets_count = export_to_excel(positions, output_path, 0, 64, 0)
                self.assertEqual(rows_count, len(positions))
                self.assertGreaterEqual(sheets_count, 1)
                workbook = load_workbook(output_path, read_only=True)
                try:
                    self.assertIn("Minerals_1", workbook.sheetnames)
                    headers = next(workbook["Minerals_1"].iter_rows(values_only=True))
                    self.assertEqual(headers[:4], ("X", "Y", "Z", "Mineral"))
                finally:
                    workbook.close()
            except Exception:
                self._write_diagnostics(server_dir, server_output)
                raise
            finally:
                self._stop_server(process, rcon)
